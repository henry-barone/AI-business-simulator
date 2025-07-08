import pandas as pd
import re
import logging
from typing import Dict, List, Optional, Union, Tuple
from pathlib import Path
import pdfplumber
import PyPDF2
from openpyxl import load_workbook
import xlrd

logger = logging.getLogger(__name__)

class PLAnalyzer:
    def __init__(self):
        self.revenue_keywords = [
            'revenue', 'sales', 'income', 'gross receipts', 'turnover',
            'total revenue', 'net sales', 'gross sales', 'operating revenue'
        ]
        
        self.cogs_keywords = [
            'cost of goods sold', 'cogs', 'cost of sales', 'cost of revenue',
            'direct costs', 'product costs', 'cost of products sold'
        ]
        
        self.labor_keywords = [
            'labor', 'wages', 'salaries', 'payroll', 'employee costs',
            'personnel costs', 'staff costs', 'compensation', 'benefits'
        ]
        
        self.overhead_keywords = [
            'overhead', 'operating expenses', 'administrative expenses',
            'general expenses', 'sg&a', 'selling general administrative',
            'general & administrative', 'general and administrative',
            'rent', 'utilities', 'insurance', 'depreciation', 'amortization'
        ]

    def parse_file(self, file_path: str) -> Dict[str, Union[float, Dict, str]]:
        file_extension = Path(file_path).suffix.lower()
        
        try:
            if file_extension in ['.xlsx', '.xls']:
                return self._parse_excel(file_path)
            elif file_extension == '.csv':
                return self._parse_csv(file_path)
            elif file_extension == '.pdf':
                return self._parse_pdf(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
                
        except Exception as e:
            logger.error(f"Error parsing file {file_path}: {str(e)}")
            return {
                'error': str(e),
                'revenue': 0.0,
                'cogs': 0.0,
                'labor_costs': 0.0,
                'overhead_costs': 0.0,
                'raw_data': {},
                'confidence_score': 0.0
            }

    def _parse_excel(self, file_path: str) -> Dict[str, Union[float, Dict, str]]:
        try:
            df = pd.read_excel(file_path, sheet_name=0)
            return self._extract_pl_data(df)
        except Exception as e:
            try:
                workbook = load_workbook(file_path)
                sheet = workbook.active
                data = []
                for row in sheet.iter_rows(values_only=True):
                    data.append(row)
                df = pd.DataFrame(data[1:], columns=data[0] if data else [])
                return self._extract_pl_data(df)
            except Exception as e2:
                logger.error(f"Failed to parse Excel file with both pandas and openpyxl: {e2}")
                raise e2

    def _parse_csv(self, file_path: str) -> Dict[str, Union[float, Dict, str]]:
        try:
            df = pd.read_csv(file_path)
            return self._extract_pl_data(df)
        except UnicodeDecodeError:
            df = pd.read_csv(file_path, encoding='latin-1')
            return self._extract_pl_data(df)
        except pd.errors.EmptyDataError:
            return {
                'revenue': 0.0,
                'cogs': 0.0,
                'labor_costs': 0.0,
                'overhead_costs': 0.0,
                'raw_data': {'dataframe_shape': (0, 0), 'columns': [], 'matched_items': [], 'sample_data': {}},
                'confidence_score': 0.0
            }

    def _parse_pdf(self, file_path: str) -> Dict[str, Union[float, Dict, str]]:
        text = ""
        
        try:
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
        except Exception as e:
            logger.warning(f"pdfplumber failed, trying PyPDF2: {e}")
            try:
                with open(file_path, 'rb') as file:
                    pdf_reader = PyPDF2.PdfReader(file)
                    for page in pdf_reader.pages:
                        text += page.extract_text() + "\n"
            except Exception as e2:
                logger.error(f"Both PDF parsers failed: {e2}")
                raise e2
        
        return self._extract_pl_from_text(text)

    def _extract_pl_data(self, df: pd.DataFrame) -> Dict[str, Union[float, Dict, str]]:
        if df.empty:
            return {
                'revenue': 0.0,
                'cogs': 0.0,
                'labor_costs': 0.0,
                'overhead_costs': 0.0,
                'raw_data': {'dataframe_shape': (0, 0), 'columns': [], 'matched_items': [], 'sample_data': {}},
                'confidence_score': 0.0
            }
        
        result = {
            'revenue': 0.0,
            'cogs': 0.0,
            'labor_costs': 0.0,
            'overhead_costs': 0.0,
            'raw_data': {},
            'confidence_score': 0.0
        }
        
        matched_items = []
        
        for idx, row in df.iterrows():
            row_str = str(row).lower()
            
            for col in df.columns:
                if pd.isna(row[col]):
                    continue
                    
                cell_value = str(row[col]).lower()
                numeric_value = self._extract_numeric_value(row[col])
                
                if numeric_value is not None and numeric_value != 0:
                    category = self._categorize_line_item(cell_value, row_str)
                    if category:
                        result[category] += abs(numeric_value)
                        matched_items.append({
                            'category': category,
                            'value': numeric_value,
                            'text': cell_value,
                            'row_index': idx,
                            'column': col
                        })
        
        result['raw_data'] = {
            'dataframe_shape': df.shape,
            'columns': list(df.columns),
            'matched_items': matched_items,
            'sample_data': df.head().to_dict() if not df.empty else {}
        }
        
        result['confidence_score'] = self._calculate_confidence_score(matched_items)
        
        return result

    def _extract_pl_from_text(self, text: str) -> Dict[str, Union[float, Dict, str]]:
        result = {
            'revenue': 0.0,
            'cogs': 0.0,
            'labor_costs': 0.0,
            'overhead_costs': 0.0,
            'raw_data': {'extracted_text': text[:1000]},
            'confidence_score': 0.0
        }
        
        lines = text.split('\n')
        matched_items = []
        
        for line_num, line in enumerate(lines):
            line_lower = line.lower().strip()
            if not line_lower:
                continue
                
            numeric_value = self._extract_numeric_value(line)
            if numeric_value is not None and numeric_value != 0:
                category = self._categorize_line_item(line_lower, line_lower)
                if category:
                    result[category] += abs(numeric_value)
                    matched_items.append({
                        'category': category,
                        'value': numeric_value,
                        'text': line.strip(),
                        'line_number': line_num
                    })
        
        result['raw_data']['matched_items'] = matched_items
        result['confidence_score'] = self._calculate_confidence_score(matched_items)
        
        return result

    def _categorize_line_item(self, text: str, context: str = "") -> Optional[str]:
        text_combined = f"{text} {context}".lower()
        
        # Check COGS first to prevent 'cost of sales' from being categorized as revenue
        if any(keyword in text_combined for keyword in self.cogs_keywords):
            return 'cogs'
        elif any(keyword in text_combined for keyword in self.revenue_keywords):
            return 'revenue'
        elif any(keyword in text_combined for keyword in self.labor_keywords):
            return 'labor_costs'
        elif any(keyword in text_combined for keyword in self.overhead_keywords):
            return 'overhead_costs'
        
        return None

    def _extract_numeric_value(self, value) -> Optional[float]:
        if pd.isna(value):
            return None
            
        if isinstance(value, (int, float)):
            return float(value)
        
        if isinstance(value, str):
            value_str = str(value).strip()
            if not value_str:
                return None
            
            # Handle parentheses for negative numbers (accounting format)
            is_negative = False
            if value_str.startswith('(') and value_str.endswith(')'):
                is_negative = True
                value_str = value_str[1:-1]
            
            # Handle currency symbols and clean the string
            value_str = re.sub(r'[$€£¥₹]', '', value_str)
            value_str = value_str.replace(',', '').replace(' ', '').lower()
            
            # Handle K/M/B suffixes
            multiplier = 1
            if value_str.endswith('k') or 'thousand' in value_str:
                multiplier = 1000
                value_str = re.sub(r'k$|thousand', '', value_str)
            elif value_str.endswith('m') or 'million' in value_str:
                multiplier = 1000000
                value_str = re.sub(r'm$|million', '', value_str)
            elif value_str.endswith('b') or 'billion' in value_str:
                multiplier = 1000000000
                value_str = re.sub(r'b$|billion', '', value_str)
            
            # Extract the numeric part
            numeric_match = re.search(r'[\d.-]+', value_str)
            if numeric_match:
                try:
                    numeric_value = float(numeric_match.group())
                    result = numeric_value * multiplier
                    return -result if is_negative else result
                except ValueError:
                    pass
        
        return None

    def _calculate_confidence_score(self, matched_items: List[Dict]) -> float:
        if not matched_items:
            return 0.0
        
        categories_found = set(item['category'] for item in matched_items)
        expected_categories = {'revenue', 'cogs', 'labor_costs', 'overhead_costs'}
        
        category_score = len(categories_found) / len(expected_categories)
        
        value_score = min(len(matched_items) / 10, 1.0)
        
        return (category_score * 0.7 + value_score * 0.3)

    def validate_data(self, parsed_data: Dict) -> Dict[str, Union[bool, List[str]]]:
        errors = []
        warnings = []
        
        if parsed_data.get('error'):
            errors.append(f"Parsing error: {parsed_data['error']}")
        
        if parsed_data.get('revenue', 0) <= 0:
            warnings.append("No revenue found or revenue is zero")
        
        if parsed_data.get('confidence_score', 0) < 0.3:
            warnings.append("Low confidence in parsed data - manual review recommended")
        
        total_costs = (parsed_data.get('cogs', 0) + 
                      parsed_data.get('labor_costs', 0) + 
                      parsed_data.get('overhead_costs', 0))
        
        if total_costs > parsed_data.get('revenue', 0) * 2:
            warnings.append("Total costs significantly exceed revenue - please verify")
        
        return {
            'is_valid': len(errors) == 0,
            'errors': errors,
            'warnings': warnings,
            'confidence_score': parsed_data.get('confidence_score', 0)
        }